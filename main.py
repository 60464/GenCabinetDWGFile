# -*- coding: utf-8 -*-
"""
time：2021-8-10
author: 60464
version: 1.1.0
"""

import os
import time
import warnings
from openpyxl import load_workbook, Workbook
from pyautocad import Autocad, APoint

warnings.filterwarnings("ignore")


def insertText(xlable, ylable, text_value, font_size):
    # 插入文字
    # 输入：x坐标，y坐标，text内容，字体大小
    insertion_pnt = APoint(xlable, ylable)
    acad.model.AddText('{0}'.format(text_value), insertion_pnt, font_size)
    print('{0} 文字已增加...'.format(text_value))


def insertMirrorText(xlable, ylable, text_value, font_size):
    # 插入镜像文字
    # 输入：x坐标，y坐标，text内容，字体大小
    insertion_pnt = APoint(xlable, ylable)
    text = acad.model.AddText('{0}'.format(text_value), insertion_pnt, font_size)
    insertion_pnt_mirror = APoint(xlable, ylable+4)
    text.Mirror(insertion_pnt, insertion_pnt_mirror)
    # 镜像后删掉原文字
    text.Delete()
    print('{0} 文字已增加...'.format(text_value))


def insertBlockFun(xlable, ylable, dwg_address, dwg_name):
    # 插入各个设备块
    # 输入：x坐标，y坐标，DWG文件夹地址，DWG文件名称（不包括后缀）
    insertion_pnt = APoint(xlable, ylable)
    dwg_address_name = dwg_address + dwg_name + '.dwg'
    if os.path.isfile(dwg_address_name):
        time.sleep(1)
        acad.model.InsertBlock(insertion_pnt, dwg_address_name, 1, 1, 1, 0)
        print('{0} 设备已增加...'.format(dwg_name))
    else:
        print(dwg_address_name, ' 不存在 ！')


def insertLoadPowerFun(xlable, ylable, dwg_address, num_value):
    # 插入各个设备块
    # 输入：x坐标，y坐标，DWG文件夹地址，DWG文件名称（不包括后缀）
    # 因为插入点在设备的左下角，当插入负载电源机笼时，需要先插入机笼，再插入1U的铭牌
    # 先判断负载电源机笼的数量，如果数量为空也是插入默认的'道岔+信号电源'
    # 如果数量=1 则需要根据选择 '道岔电源', '信号电源', '道岔+信号电源' 三种的一种
    # 如果数量大于1 则直接选择'道岔+信号电源'这一种，因为既然选择多个 肯定是每一个都是满配
    if num_value is None:
        num_value = 1
    if num_value == 1 :
        if board_name_value in ['道岔电源', '信号电源', '道岔+信号电源']:
            # 更新y轴坐标 插入1U铭牌
            ylable = ylable + one_u_high * 3
            insertBlockFun(xlable, ylable, dwg_address, '负载配电机箱-铭牌')
            ylable = ylable - one_u_high * 3
            # 插入机笼
            insertBlockFun(xlable, ylable, dwg_address, '负载配电机箱-'+board_name_value)
    else:
        # 超过两个的情况
        # 更新y轴坐标 插入1U铭牌
        ylable = ylable + one_u_high * 3
        insertBlockFun(xlable, ylable, dwg_address, '负载配电机箱-铭牌')
        for i in range(num_value):
            # 插入机笼
            ylable = ylable - one_u_high * 3
            insertBlockFun(xlable, ylable, dwg_address, '负载配电机箱-道岔+信号电源')

    return ylable


def insertViopBoardFun(num_value, xlable_initial, io_with_com_a_x, io_board_thicknss_offset, ylable,
                       io_with_com_a_y, io_with_com_b_x, io_with_com_b_y,
                       dwg_address, board_name_value):
    for k in range(num_value):
        # 画A机板块
        insertBlockFun(xlable_initial + io_with_com_a_x +
                       k * io_board_thicknss_offset, ylable +
                       io_with_com_a_y, dwg_address, board_name_value)
        # 画B机板块
        insertBlockFun(xlable_initial + io_with_com_b_x +
                       k * io_board_thicknss_offset, ylable +
                       io_with_com_b_y, dwg_address, board_name_value)


def insertEcidBoardFun(num_value, xlable_initial, ecid_with_com_a_x, ecid_board_thicknss_offset,
                       ylable, ecid_with_com_a_y, ecid_with_com_b_x, ecid_with_com_b_y,
                       dwg_address, board_name_value):
    for k in range(num_value):
        # 画A机板块
        insertBlockFun(
            xlable_initial + ecid_with_com_a_x +
            k * ecid_board_thicknss_offset,
            ylable + ecid_with_com_a_y, dwg_address, board_name_value)
        # 画B机板块
        insertBlockFun(
            xlable_initial + ecid_with_com_b_x +
            k * ecid_board_thicknss_offset,
            ylable + ecid_with_com_b_y, dwg_address, board_name_value)

def summaryJiguiEquipment(ips_jigui_dic, equipments_dic):
    # 本函数用于汇总每个机柜内的设备数量
    ips_jigui_total_dic = {}
    for name, value in ips_jigui_dic.items():
        # 获取传递来的设备名称和设备数量
        if equipments_dic.get(name, 0) != 0:
            for name_single, value_single in equipments_dic[name].items():
                # 获取equipments_dic 每个单项的名称和数量
                # print(name_single, value_single)
                if isinstance(value_single, int):
                    # 如果value是int类型，则判断ips_jigui_total_dic是否已经有了，根据是否有了对数值进行增加
                    if ips_jigui_total_dic.get(name_single, 0) != 0:
                        # 如果存在，则更新数量
                        ips_jigui_total_dic[name_single] = ips_jigui_total_dic[name_single] + value_single * value
                    else:
                        # 如果不存在，则直接赋值
                        ips_jigui_total_dic[name_single] = value_single * value
                if isinstance(value_single, dict):
                    # 如果value是字典类型，则需要分解字典
                    # print('name and value:',name_single, value_single)
                    # print('total_dic',ips_jigui_total_dic)
                    if ips_jigui_total_dic.get(name_single, 0) != 0:
                        # 如果存在，则更新数量
                        ips_jigui_total_dic[name_single]['Number'] = ips_jigui_total_dic[name_single]['Number'] + \
                                                                     value_single['Number'] * value
                        ips_jigui_total_dic[name_single]['To'] = ips_jigui_total_dic[name_single]['To'] + \
                                                                 value_single['To'] * value
                    else:
                        # 如果不存在，则直接增加到字典中
                        ips_jigui_total_dic.update({name_single: {'Number': value_single['Number'] * value,
                                                                  'To': value_single['To'] * value
                                                                  }
                                                    })
        else:
            continue
    return ips_jigui_total_dic


def equipmentsListExecl(ips_jigui_dic, execute_jigui_dic, station_name):
    # 生成设备清单，分执行机柜部分的设备和联锁机柜的设备清单
    # 机柜内部配线，比如铜条就不需要了
    # 定义设备字典，key是设备名称，value是对应的电缆数量 或者其他设备数量
    equipments_dic = {
        # CVC配套的是：电源线，串口线，CAN线，网线
        'CVCA机笼': {'CVC机笼': 1, 'PSU板': 2, 'CMU板': 1, 'MLU板': 4,
                    'CD0065线缆': {
                                    'Number': 6,
                                    'To':
                                        [
                                            'A-HSSL_1 -> B-HSSL_1',
                                            'A-HSSL_2 -> B-HSSL_2',
                                            'A-MPU1_N4 -> B-MPU1_N4',
                                            'A-MPU1_N3 -> B-MPU1_N3',
                                            'A-MPU2_N4 -> B-MPU2_N4',
                                            'A-MPU2_N3 -> B-MPU2_N3'
                                        ]
                                  },
                    'CD125线缆': {
                                    'Number': 2,
                                    'To':
                                        ['A-MCU_C1 -> STBY-H-CAN_RS422_A1', 'A-MCU_C2 -> STBY-H-CAN_RS422_B1']
                                 },
                    'CD0063线缆': {
                                    'Number': 8,
                                    'To':
                                        [
                                            'A-MCU_N1  ->  SWA-1',
                                            'A-MCU_N2  ->  SWB-1',
                                            'A-MNCU1_N1 -> SWA-3',
                                            'A-MNCU1_N2 -> SWB-3',
                                            'A-MNCU2_N1 -> SWA-5',
                                            'A-MNCU2_N2 -> SWB-5',
                                            'A-MNCU1_N3 -> SWA-7',
                                            'A-MNCU1_N4 -> SWB-7'
                                        ]
                                },
                    'CD006线缆': {'Number': 2, 'To': ['A-联锁机pwr1 -> PDU-A11', 'A-联锁机pwr2 -> PDU-A12']}
                   },
        'CVCB机笼': {'CVC机笼': 1, 'PSU板': 2, 'CMU板': 1, 'MLU板': 4,
                    'CD125线缆': {
                                    'Number': 2,
                                    'To':
                                        ['B-MCU_C1 -> STBY-H-CAN_RS422_A1', 'B-MCU_C2 -> STBY-H-CAN_RS422_B1']
                                 },
                    'CD0063线缆': {
                                    'Number': 8,
                                     'To': [
                                           'B-MCU_N1  ->  SWA-2',
                                           'B-MCU_N2  ->  SWB-2',
                                           'B-MNCU1_N1 -> SWA-4',
                                           'B-MNCU1_N2 -> SWB-4',
                                           'B-MNCU2_N1 -> SWA-6',
                                           'B-MNCU2_N2 -> SWB-6',
                                           'B-MNCU1_N3 -> SWA-8',
                                           'B-MNCU1_N4 -> SWB-8'
                                            ]
                                },
                    'CD006线缆': {'Number': 2, 'To': ['B-联锁机pwr1 -> PDU-B11', 'B-联锁机pwr2 -> PDU-B12']}
                   },
        # STBY配套的是：
        'STBY主备切换单元': {
                           'STBY主备切换单元': 1,
                           'CD006线缆': {'Number': 2, 'To': ['STBY-pwr1 -> PDU-B11', 'STBY-pwr2 -> PDU-B12']}
                          },
        # 交换机配套的是：电源线
        '交换机A': {'交换机': 1, 'CD006线缆': {'Number': 1, 'To': ['A-交换机 -> PDU-A21']}},
        '交换机B': {'交换机': 1, 'CD006线缆': {'Number': 1, 'To': ['B-交换机 -> PDU-B21']}},
        # 工控机配套的是：电源线，网线，（有KVM不需要网线）
        '操作机A': {'KA8005（包括键鼠）': 1, '显示器': 1, '音箱': 1,
                  'CD006线缆': {'Number': 1, 'To': ['MMIA-PWR -> PDU-A11']},
                  '屏蔽网线': {'Number': 2, 'To': ['MMIA-NETA -> SWA-10', 'MMIA-NETB -> SWB-10']}
                  },
        '操作机B': {'KA8005（包括键鼠）': 1, '显示器': 1, '音箱': 1,
                 'CD006线缆': {'Number': 1, 'To': ['MMIB-PWR -> PDU-B11']},
                 '屏蔽网线': {'Number': 2, 'To': ['MMIB-NETA -> SWA-11', 'MMIB-NETB -> SWB-11']}
                 },
        # kvm配套的是：电源线
        'KVM': {'STBY主备切换单元': 1, '电源线': 2},
        # 工控机配套的是：电源线，网线，串口线
        '维修机显示器': {'维修机显示器': 1, '电源线': 2},
        '维修机': {'KA8005（包括键鼠）': 1,
                 'CD006线缆': {'Number': 1, 'To': ['SDM-PWR -> PDU-A11']},
                 '屏蔽网线': {'Number': 2, 'To': ['SDM-NETA -> SWA-12', 'SDM-NETB -> SWB-12']}},
        # 直流电源显示单元配套的是：电源线，串口线
        '直流电源显示单元': {'CD0069': 1, 'CD0069-2': 1, 'CD0069-3': 1},
        'UPSA': {'UPS': 1},
        'UPSB': {'UPS': 1},
        '交流电源切换单元A': {'交流电源切换单元': 1},
        '交流电源切换单元B': {'交流电源切换单元': 1},
        '配电箱': {'配电箱': 1},
        # 执行单元(带COM)配套的是：机笼，EIOCOM4板，两个，网线1个，CAN线是个问题
        '执行单元(带COM)': {'带COM机笼': 1, 'BP5母板': 1, 'EIOCOM4板': 2,
                          'CD107-3线缆': {'Number': 1, 'To': ['EIOCOM4 -> SWA-15,SWB-15']},
                          'CD125线缆': 2},
        '执行单元(不带COM)': {'不带COM机笼': 1, 'BP5母板': 1, 'CD125线缆': 2},
        '电源机箱': {'电源机笼': 1, '电源模块': 2,
                    'CD006线缆': {'Number': 2, 'To': ['A-PSU -> PDU-A51', 'B-PSU -> PDU-B51']}
                   },
        # 传递过来的板卡上数量是值的单系的，所以这里的字典，板卡数量是2，线缆是1（ECID）或2（VIOP）
        'VIIB32S-2': {'VIIB32S-2板': 2, 'CD403线缆': 2},
        'VIIB32D-2': {'VIIB32S-2板': 2, 'CD405线缆': 2},
        'VOOB16-2': {'VOOB16-2板': 2, 'CD404线缆': 2},
        'SDDM-2': {'SDDM-2板': 2, 'CD300线缆': 1},
        'PDDM5-2': {'PDDM5-2板': 2, 'CD306线缆': 1},
        'PDDM46-2': {'PDDM46-2板': 2, 'CD302线缆': 1},
        'TCIM-2': {'TCIM-2板': 2, 'CD303线缆': 1},
        'SIOM-2': {'SIOM-2板': 2, 'CD302线缆': 1},
        'CDDM-2': {'CDDM-2板': 2, 'CDXXX线缆': 1},
        'HIOM-2': {'HIOM-2板': 2, 'CD308线缆': 1}
    }
    ips_jigui_dic_flag = 1
    execute_jigui_dic_flag = 1
    ips_jigui_total_dic = {}
    execute_jigui_total_dic = {}
    # 分别计算综合机柜和执行机柜的设备数量
    if len(ips_jigui_dic) > 0:
        # 只有传进来的不为空才执行
        ips_jigui_total_dic = summaryJiguiEquipment(ips_jigui_dic, equipments_dic)
        print(ips_jigui_total_dic)
        ips_jigui_dic_flag = 0

    if len(execute_jigui_dic) > 0:
        # 只有传进来的不为空才执行
        execute_jigui_total_dic = summaryJiguiEquipment(execute_jigui_dic, equipments_dic)
        print(execute_jigui_total_dic)
        execute_jigui_dic_flag = 0

    # 写execl
    # 建一个.xlsx
    wb = Workbook()
    # workbook在创建的时候同时至少也新建了一张工作表(worksheet)。通过openpyxl.workbook.Workbook.active()调用得到正在运行的工作表。
    # ws = wb.active  # ws操作sheet页
    sheet = wb.create_sheet('Equipments List', 0)
    sheet['A1'] = '序号'
    sheet['B1'] = '设备名称'
    sheet['C1'] = '数量'
    sheet['D1'] = 'From->To'
    sheet['E1'] = '长度'
    sheet['F1'] = '备注'

    if ips_jigui_dic_flag == 0:
        # 有联锁综合柜则执行如下内容
        sheet['A2'] = '联锁机柜设备'
        # gap_in_ips_jigui初始=3，是因为是从表格的第三行开始写
        gap_in_ips_jigui = 3
        for i, (name, value) in enumerate(ips_jigui_total_dic.items()):
            try:
                sheet['A' + str(i + gap_in_ips_jigui)] = i+1
                sheet['B' + str(i + gap_in_ips_jigui)] = name
                if isinstance(value, int):
                    sheet['C' + str(i + gap_in_ips_jigui)] = value
                if isinstance(value, dict):
                    sheet['C' + str(i + gap_in_ips_jigui)] = value['Number']
                    for i_2, from_to_item in enumerate(value['To']):
                        # 线缆单独的作为一行表格
                        sheet['D' + str(i + gap_in_ips_jigui + i_2)] = from_to_item
                    gap_in_ips_jigui = gap_in_ips_jigui + len(value['To']) - 1
            except:
                continue
        gap_in_execute_jigui = i + gap_in_ips_jigui + 1
    if execute_jigui_dic_flag == 0:
        # 有联锁综合柜则执行如下内容
        sheet['A'+str(gap_in_execute_jigui)] = '执行机柜设备'
        for j, (name, value) in enumerate(execute_jigui_total_dic.items()):
            try:
                sheet['A' + str(j + gap_in_execute_jigui + 1)] = j+1
                sheet['B' + str(j + gap_in_execute_jigui + 1)] = name
                if isinstance(value, int):
                    sheet['C' + str(j + gap_in_execute_jigui + 1)] = value
                if isinstance(value, dict):
                    sheet['C' + str(j + gap_in_execute_jigui + 1)] = value['Number']
                    for j_2, from_to_item in enumerate(value['To']):
                        # 线缆单独的作为一行表格
                        sheet['D' + str(j + gap_in_execute_jigui + 1 + j_2)] = from_to_item
                    gap_in_execute_jigui = gap_in_execute_jigui + len(value['To']) - 1
            except:
                continue

    add_hour = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()).split()[1].replace(':', '_')
    wb.save('Equipments_List_{0}_{1}.xlsx'.format(station_name, add_hour))

    return


# 获取当前地址并定义相关文件名称
current_path = os.getcwd()
dwg_folder = '\\DWG\\'
dwg_address = current_path + dwg_folder
config_name = 'Config_.xlsx'
folder_name = dwg_folder[1:4]

# 判断文件夹和文件是否存在
print('==============================================================')
dwg_flag = 0
excel_flag = 0
if os.path.exists(folder_name):
    # 存在文件夹，继续进行
    dwg_flag = 1
else:
    print('DWG文件夹不存在，请拷贝该文件夹及内容后再次执行本程序')
files = os.listdir(current_path)
for item in files:
    if 'Config_' and '.xlsx' in item:
        # 存在该文件，继续执行
        excel_flag = 1
        config_name = item
        station_name = config_name.split('_')[1].split('.')[0]
        break
else:
    print('Config_???.xlsx文件不存在，请拷贝该文件后再次执行本程序')

# 对 dwg_flag 和 excel_flag 进行判断
if dwg_flag * excel_flag == 1:
    print('请确保AutoCAD软件已打开')
    # ==============定义相关坐标
    # 如下两个坐标用于机柜定位（机柜左下角）
    jigui_xlable = 54
    jigui_ylable = 35
    # 如下两个坐标用于机柜内部设备定位（机柜左上角）
    jigui_49u_x = 4.75
    jigui_49u_y = 229.912
    jigui_47u_x = 4.75
    jigui_47u_y = 220.88
    one_u_high = 4.45
    # 如下两个参数用于定位设备名称文字及其下划线
    jigui_thicknss_offset = 5
    text_up_offset = 4
    line_length = 10
    part_line_length = 6.9677
    # 机柜最上面名称的文字坐标
    jigui_name_xlable = 10
    jigui_name_ylable = 10
    # 设备名称字体大小
    name_font_size = 1.5
    # 两个机柜之间距离（两个左下角之间的距离）
    gap_between_jigui = 80
    # 如下几个参数用于ECID机笼内容定位板卡的坐标
    ecid_with_com_a_x = 6.8842
    ecid_with_com_a_y = 0.2126
    ecid_with_com_b_x = 29.2499
    ecid_with_com_b_y = ecid_with_com_a_y
    ecid_board_thicknss_offset = 10.9507-6.8842
    ecid_without_com_a_x = ecid_with_com_a_x - ecid_board_thicknss_offset
    ecid_without_com_a_y = ecid_with_com_a_y
    ecid_without_com_b_x = ecid_with_com_b_x - ecid_board_thicknss_offset
    ecid_without_com_b_y = ecid_with_com_a_y
    # 如下几个参数用于IO机笼内容定位板卡的坐标
    io_with_com_a_x = 7.8862
    io_with_com_a_y = 0.2621
    io_with_com_b_x = 27.6665
    io_with_com_b_y = ecid_with_com_a_y
    io_board_thicknss_offset = 10.4222-7.8862
    io_without_com_a_x = io_with_com_a_x - io_board_thicknss_offset
    io_without_com_a_y = ecid_with_com_a_y
    io_without_com_b_x = io_with_com_b_x - io_board_thicknss_offset
    io_without_com_b_y = ecid_with_com_a_y
    # 执行机箱内部的编号的坐标
    jixiang_first_row_text_xlable = 27.00
    jixiang_first_row_text_ylable = 29.2138
    jixiang_first_row_text_high = 0.8

    # 定义设备清单list
    ips_jigui_dic = {}
    execute_jigui_dic = {}

    # 读取execl表格，先判断execl表格里面是否有sheet，如果有则进行画图
    # 读取配置文件config.xlsx
    excel = load_workbook(config_name, data_only=True)
    # 获取所有表格名称
    sheets = excel.sheetnames
    ips_jigui_flag = 0
    execute_jigui_flag = 0
    # 判断是否存在联锁机机柜和执行机机柜
    execute_jigui_name = []
    for sheet_name in sheets:
        if '联锁机机柜' in sheet_name:
            ips_jigui_flag = 1
        if '执行机柜' in sheet_name:
            execute_jigui_name.append(sheet_name)
            execute_jigui_flag = 1
    # 只要有联锁机机柜或者执行机机柜就可以执行
    if ips_jigui_flag == 1 or execute_jigui_flag == 1:
        # 连接AutoCAD
        acad = Autocad(create_if_not_exists=True)
        # 新建一个文件,无法直接命名
        acad.Application.Documents.Add('')
        try:
            # 设置文字的字体样式，将特定文字样式设为当前
            acad.ActiveDocument.ActiveTextStyle = acad.ActiveDocument.TextStyles.Item("Standard")
            # 为当前文本样式设置字体
            acad.ActiveDocument.ActiveTextStyle.SetFont("微软雅黑", False, False, 1, 0 or 0)
        except:
            print('字体设置有问题，但不影响机柜生成！！')

        # 插入图纸标准外框
        insertBlockFun(0, 0, dwg_address, '图纸标准外框')
        # 联锁机机柜配置 sheet
        if '联锁机机柜配置' in sheets:
            table = excel['联锁机机柜配置']
            # 读取机柜尺寸，并画图
            jigui_heigh = table['C3'].value
            # 插入机柜框
            insertBlockFun(jigui_xlable, jigui_ylable, dwg_address, jigui_heigh)
            # 判断机柜高度，确定设备坐标原点
            # 除非是选择47U，不选择单元格内容或者选择49U,机柜高度都是49U
            xlable_initial = jigui_xlable + jigui_49u_x
            ylable_initial = jigui_ylable + jigui_49u_y
            if jigui_heigh == '47U':
                xlable_initial = jigui_xlable + jigui_47u_x
                ylable_initial = jigui_ylable + jigui_47u_y
            # 增加机柜名称text
            text_value = table.cell(row=2, column=1).value
            insertText(xlable_initial+jigui_name_xlable, ylable_initial+jigui_name_ylable, text_value, 4)
            # 画机柜内部的机笼
            # 首先判断一下 最后一行序号是多少
            last_rows = table.max_row
            for i in range(5, last_rows):
                cell_value = table['B' + str(i)].value
                high_value = table['C' + str(i)].value
                flag_value = table['D' + str(i)].value
                text_value = table['E' + str(i)].value
                # print(cell_value,high_value,flag_value)
                try:
                    if flag_value == '是':
                        # 如果 cell_value 或者 high_value为空 则这一层就不画了，位置也不更新
                        if cell_value is None or high_value is None:
                            continue
                        # 如果选择配置，则执行
                        if i != 5:
                            # 增加设备列表
                            # 先判断字典中是否存在该设备，如果存在，则数目增加，如果不存在，直接赋值数目
                            if ips_jigui_dic.get(cell_value, 0) != 0:
                                ips_jigui_dic[cell_value] = ips_jigui_dic[cell_value] + 1
                            else:
                                ips_jigui_dic[cell_value] = 1
                        # 画每一行的设备
                        ylable = ylable_initial - high_value * one_u_high
                        # 插入每一行设备
                        insertBlockFun(xlable_initial, ylable, dwg_address, cell_value)
                        # 更新y轴坐标
                        ylable_initial = ylable
                        # 画每一个设备对应的名称
                        if text_value is not None:
                            # 如果设备对应名称列有内容，则画设备名称下划线和写名称内容
                            line_ylable = ylable + 0.5 * high_value * one_u_high
                            line_xlable = xlable_initial - line_length - jigui_thicknss_offset
                            # 插入下划线
                            insertBlockFun(line_xlable, line_ylable, dwg_address, '设备名称下划线')
                            # 插入镜像文字
                            insertMirrorText(line_xlable + part_line_length, line_ylable + text_up_offset, text_value,
                                             name_font_size)
                    else:
                        # 如果选择否，或者空白内容，则填充补空板
                        for m in range(high_value):
                            ylable = ylable_initial - one_u_high
                            # 插入1u补空板
                            insertBlockFun(xlable_initial, ylable, dwg_address, '1U补空板')
                            # 更新y轴坐标
                            ylable_initial = ylable
                except:
                    continue
            print('联锁机机柜生成完毕')
        else:
            print('联锁机机柜配置sheet不存在!')
        # 执行机柜配置 sheet
        if len(execute_jigui_name) > 0:
            for i, execute_sheet_name in enumerate(execute_jigui_name):
                table = excel[execute_sheet_name]
                # 读取机柜尺寸，并画图
                jigui_heigh = table['C3'].value
                insertBlockFun(jigui_xlable + (i+1) * gap_between_jigui, jigui_ylable, dwg_address, jigui_heigh)
                # 判断机柜高度，确定设备坐标原点
                # 由于已经增加国其他机柜，所以需要增加机柜间距离的GAP
                # 除非是选择47U，不选择单元格内容或者选择49U,机柜高度都是49U
                xlable_initial = jigui_xlable + (i+1) * gap_between_jigui + jigui_49u_x
                ylable_initial = jigui_ylable + jigui_49u_y
                if jigui_heigh == '47U':
                    xlable_initial = jigui_xlable + (i+1) * gap_between_jigui + jigui_47u_x
                    ylable_initial = jigui_ylable + jigui_47u_y
                # 增加机柜名称text
                text_value = table.cell(row=2, column=1).value
                insertText(xlable_initial + jigui_name_xlable, ylable_initial + jigui_name_ylable, text_value, 4)
                # 首先判断一下 最后一行序号是多少
                last_rows = table.max_row
                # 画采驱机柜设备
                for j in range(5, last_rows):
                    try:
                        cell_value = table['B' + str(j)].value
                        high_value = table['C' + str(j)].value
                        flag_value = table['D' + str(j)].value
                        text_value = table['G' + str(j)].value
                        num_value = table['F' + str(j)].value
                        board_name_value = table['E' + str(j)].value
                        if flag_value == '是':
                            # 如果cell_value为空 high_value肯定也是空的（execl自动生成）
                            # 这个时候直接跳过这一层，下一层将覆盖这一层
                            if cell_value is None or high_value is None:
                                continue
                            # 画每一行的设备
                            ylable = ylable_initial - high_value * one_u_high
                            # 如果是板块机笼，则先判断 num_value 和 board_name_value num_value是否为空，为空就不执行后面了，机柜图中这曾也空着
                            if cell_value in ['执行单元(带COM)', '执行单元(不带COM)']:
                                if cell_value is None or num_value is None or board_name_value is None:
                                    # 虽然不执行画图操作，但是坐标还是要更新，把这部分空出来
                                    ylable_initial = ylable
                                    continue

                            if j != 5:
                                # 增加设备列表
                                # 先判断字典中是否存在该设备，如果存在，则数目增加，如果不存在，直接赋值数目
                                if execute_jigui_dic.get(cell_value, 0) != 0:
                                    execute_jigui_dic[cell_value] = execute_jigui_dic[cell_value] + 1
                                else:
                                    execute_jigui_dic[cell_value] = 1
                                # 判断字典是否存在板卡,当EXCEL中板块名称存在且数量不=0的时候
                                if board_name_value is not None and num_value != 0:
                                    if execute_jigui_dic.get(board_name_value, 0) != 0:
                                        execute_jigui_dic[board_name_value] = execute_jigui_dic[board_name_value] + \
                                                                              num_value
                                    else:
                                        execute_jigui_dic[board_name_value] = num_value

                            if cell_value in ['执行机柜铭牌', '电源机箱']:
                                insertBlockFun(xlable_initial, ylable, dwg_address, cell_value)
                            elif cell_value == '负载配电机箱':
                                # 如果是负载配电机箱，则单独执行负载配电机箱的函数
                                ylable = insertLoadPowerFun(xlable_initial, ylable, dwg_address, num_value)
                                # 更新y坐标
                                ylable_initial = ylable
                            elif cell_value in ['执行单元(带COM)', '执行单元(不带COM)']:
                                # 从第六行至倒数第二行开始，并开始判断板卡名称
                                dwg_jilong_name = dwg_address + cell_value + '.dwg'
                                # dwg_board_name = dwg_address + board_name_value + '.dwg'
                                if os.path.isfile(dwg_jilong_name):
                                    # 先判断是带COM的还是不带COM的
                                    if cell_value == '执行单元(带COM)':
                                        # 如果是采集驱动板，则需要采用IO模式的机笼
                                        if board_name_value in ['VIIB32S-2', 'VIIB32D-2', 'VOOB16-2', 'VIIB32S-2&VOOB16-2', 'VIIB32D-2&VOOB16-2']:
                                            # 插入机笼
                                            insertBlockFun(xlable_initial, ylable, dwg_address, '执行单元(带COM)_IO')
                                            # 插入机笼的序号
                                            insertText(xlable_initial + jixiang_first_row_text_xlable,
                                                       ylable + jixiang_first_row_text_ylable, str(j-5),
                                                       jixiang_first_row_text_high)
                                            # 画板卡
                                            # 画板卡前先判断是否是VIOP的混插，如果是VIOP的混插 就不检查 num_value是否为空 而是检查 I J K L四列的情况
                                            if board_name_value in ['VIIB32S-2&VOOB16-2', 'VIIB32D-2&VOOB16-2']:
                                                # 获取混插板卡数量
                                                board_extend1_name_value = table['I' + str(j)].value
                                                board_extend1_num_value = table['J' + str(j)].value
                                                board_extend2_name_value = table['K' + str(j)].value
                                                board_extend2_num_value = table['L' + str(j)].value
                                                # 判断混插板卡的名称和数量, 不为空的时候再插入
                                                if board_extend1_name_value is not None and board_extend1_num_value is not None:
                                                    board_extend1_num_value = min(board_extend1_num_value, 6)
                                                    insertViopBoardFun(board_extend1_num_value, xlable_initial, io_with_com_a_x,
                                                                       io_board_thicknss_offset, ylable,
                                                                       io_with_com_a_y, io_with_com_b_x,
                                                                       io_with_com_b_y,
                                                                       dwg_address, board_extend1_name_value)
                                                # 前面画的几个板子后 后面再画板卡要加上之前这个的距离
                                                board_extend2_gap = io_board_thicknss_offset * board_extend1_num_value
                                                if board_extend2_name_value is not None and board_extend2_num_value is not None:
                                                    board_extend2_num_value = min(board_extend2_num_value, 6)
                                                    insertViopBoardFun(board_extend2_num_value, xlable_initial + board_extend2_gap, io_with_com_a_x,
                                                                       io_board_thicknss_offset, ylable,
                                                                       io_with_com_a_y, io_with_com_b_x,
                                                                       io_with_com_b_y,
                                                                       dwg_address, board_extend2_name_value)
                                            else:
                                                # 不是VIOP的混插 则需要判断 num_value 是否为空
                                                if num_value is not None:
                                                    # 对板块数量进行判断，带com的不能超过6个
                                                    num_value = min(num_value, 6)
                                                    insertViopBoardFun(num_value, xlable_initial, io_with_com_a_x,
                                                                       io_board_thicknss_offset, ylable,
                                                                       io_with_com_a_y, io_with_com_b_x,
                                                                       io_with_com_b_y,
                                                                       dwg_address, board_name_value)

                                        else:
                                            insertBlockFun(xlable_initial, ylable, dwg_address, cell_value)
                                            insertText(xlable_initial + jixiang_first_row_text_xlable,
                                                       ylable + jixiang_first_row_text_ylable, str(j - 5),
                                                       jixiang_first_row_text_high)
                                            # 画板卡
                                            if board_name_value == '多板卡混用':
                                                # 获取混插板卡数量
                                                for board_ex in range(4):
                                                    board_extend_name_value = table[chr(ord('I') + 2 * board_ex ) + str(j)].value
                                                    board_extend_num_value = table[
                                                        chr(ord('I') + 1 + 2 * board_ex ) + str(j)].value
                                                    # 判断混插板卡的名称和数量, 不为空的时候再插入
                                                    if board_extend_name_value is not None and board_extend_num_value is not None:
                                                        board_extend_num_value = min(board_extend_num_value, 6)
                                                        insertEcidBoardFun(board_extend_num_value, xlable_initial, ecid_with_com_a_x,
                                                                           ecid_board_thicknss_offset,
                                                                           ylable, ecid_with_com_a_y, ecid_with_com_b_x,
                                                                           ecid_with_com_b_y,
                                                                           dwg_address, board_extend_name_value)
                                                        # 更新画下一个板卡的x轴距离
                                                        xlable_initial = xlable_initial + board_extend_num_value * ecid_board_thicknss_offset
                                            else:
                                                if num_value is not None:
                                                    # 对板块数量进行判断，带com的不能超过4个
                                                    num_value = min(num_value, 4)
                                                    insertEcidBoardFun(num_value, xlable_initial, ecid_with_com_a_x,
                                                                       ecid_board_thicknss_offset,
                                                                       ylable, ecid_with_com_a_y, ecid_with_com_b_x,
                                                                       ecid_with_com_b_y,
                                                                       dwg_address, board_name_value)

                                    if cell_value == '执行单元(不带COM)':
                                        # 如果是采集驱动板，则需要采用IO模式的机笼
                                        if board_name_value in ['VIIB32S-2', 'VIIB32D-2', 'VOOB16-2', 'VIIB32S-2&VOOB16-2', 'VIIB32D-2&VOOB16-2']:
                                            insertBlockFun(xlable_initial, ylable, dwg_address, '执行单元(不带COM)_IO')
                                            insertText(xlable_initial + jixiang_first_row_text_xlable,
                                                       ylable + jixiang_first_row_text_ylable, str(j - 5),
                                                       jixiang_first_row_text_high)
                                            # 画板卡
                                            # 画板卡前先判断是否是VIOP的混插，如果是VIOP的混插 就不检查 num_value是否为空 而是检查 I J K L四列的情况
                                            if board_name_value in ['VIIB32S-2&VOOB16-2', 'VIIB32D-2&VOOB16-2']:
                                                # 获取混插板卡数量
                                                board_extend1_name_value = table['I' + str(j)].value
                                                board_extend1_num_value = table['J' + str(j)].value
                                                board_extend2_name_value = table['K' + str(j)].value
                                                board_extend2_num_value = table['L' + str(j)].value
                                                # 判断混插板卡的名称和数量, 不为空的时候再插入
                                                if board_extend1_name_value is not None and board_extend1_num_value is not None:
                                                    board_extend1_num_value = min(board_extend1_num_value, 7)
                                                    insertViopBoardFun(board_extend1_num_value, xlable_initial,
                                                                       io_without_com_a_x,
                                                                       io_board_thicknss_offset, ylable,
                                                                       io_without_com_a_y, dwg_address,
                                                                       board_extend1_name_value)
                                                # 前面画的几个板子后 后面再画板卡要加上之前这个的距离
                                                board_extend2_gap = io_board_thicknss_offset * board_extend1_num_value
                                                if board_extend2_name_value is not None and board_extend2_num_value is not None:
                                                    board_extend2_num_value = min(board_extend2_num_value, 7)
                                                    insertViopBoardFun(board_extend2_num_value, xlable_initial + board_extend2_gap,
                                                                       io_without_com_a_x,
                                                                       io_board_thicknss_offset, ylable,
                                                                       io_without_com_a_y, dwg_address,
                                                                       board_extend2_name_value)
                                            else:
                                                if num_value is not None:
                                                    # 对板块数量进行判断，带com的不能超过6个
                                                    num_value = min(num_value, 7)
                                                    insertViopBoardFun(num_value, xlable_initial,
                                                                       io_without_com_a_x,
                                                                       io_board_thicknss_offset, ylable,
                                                                       io_without_com_a_y, dwg_address,
                                                                       board_name_value)

                                        else:
                                            insertBlockFun(xlable_initial, ylable, dwg_address, cell_value)
                                            insertText(xlable_initial + jixiang_first_row_text_xlable,
                                                       ylable + jixiang_first_row_text_ylable, str(j - 5),
                                                       jixiang_first_row_text_high)
                                            # 画板卡
                                            if board_name_value == '多板卡混用':
                                                # xlable_tmp是用于多板卡混插时，每个板卡插入的x坐标都是基于前面已经插入的位置
                                                xlable_tmp = 0
                                                for board_ex in range(5):
                                                    board_extend_name_value = table[chr(ord('I') + 2 * board_ex) + str(j)].value
                                                    board_extend_num_value = table[
                                                        chr(ord('I') + 2 * board_ex + 1) + str(j)].value
                                                    # print(board_extend_name_value, board_extend_num_value)
                                                    # 判断混插板卡的名称和数量, 不为空的时候再插入
                                                    if board_extend_name_value is not None and board_extend_num_value is not None:
                                                        board_extend_num_value = min(board_extend_num_value, 7)
                                                        insertEcidBoardFun(board_extend_num_value, xlable_initial + xlable_tmp, ecid_without_com_a_x,
                                                                           ecid_board_thicknss_offset,
                                                                           ylable, ecid_without_com_a_y, ecid_without_com_b_x,
                                                                           ecid_without_com_b_y,
                                                                           dwg_address, board_extend_name_value)
                                                        # 更新画下一个板卡的x轴距离
                                                        xlable_tmp = xlable_tmp + board_extend_num_value * ecid_board_thicknss_offset
                                            else:
                                                if num_value is not None:
                                                    # 对板块数量进行判断，带com的不能超过4个
                                                    num_value = min(num_value, 5)
                                                    insertEcidBoardFun(num_value, xlable_initial, ecid_without_com_a_x,
                                                                       ecid_board_thicknss_offset,
                                                                       ylable, ecid_without_com_a_y, ecid_without_com_b_x,
                                                                       ecid_without_com_b_y,
                                                                       dwg_address, board_name_value)

                                else:
                                    print('{0}不存在 ！'.format(dwg_jilong_name))
                            else:
                                insertBlockFun(xlable_initial, ylable, dwg_address, cell_value)
                            ylable_initial = ylable
                            # 画每一个设备对应的名称
                            if text_value is not None:
                                # 如果设备对应名称列有内容，则画设备名称下划线和写名称内容
                                line_ylable = ylable + 0.5 * high_value * one_u_high
                                line_xlable = xlable_initial - line_length - jigui_thicknss_offset
                                # 插入下划线
                                insertBlockFun(line_xlable, line_ylable, dwg_address, '设备名称下划线')
                                # 插入镜像文字
                                insertMirrorText(line_xlable + part_line_length, line_ylable + text_up_offset,
                                                 text_value, name_font_size)

                        else:
                            # 如果选择否，或者空白内容，则填充补空板
                            if high_value is not None:
                                # 先判断尺寸单元格是否有值，没有则直接跳过补空板的增加
                                if high_value == 7:
                                    # 如果是7U的机笼，只要补空1U+6U就行
                                    # 插入1u补空板
                                    ylable = ylable_initial - one_u_high
                                    insertBlockFun(xlable_initial, ylable, dwg_address, '1U补空板')
                                    # 更新y轴坐标
                                    ylable_initial = ylable
                                    # 插入6u补空板
                                    ylable = ylable_initial - 6 * one_u_high
                                    insertBlockFun(xlable_initial, ylable, dwg_address, '6U补空板')
                                    # 更新y轴坐标
                                    ylable_initial = ylable
                                elif high_value == 5:
                                    # 如果是5U的机笼，只要补空1U+4U就行
                                    # 插入1u补空板
                                    ylable = ylable_initial - one_u_high
                                    insertBlockFun(xlable_initial, ylable, dwg_address, '1U补空板')
                                    # 更新y轴坐标
                                    ylable_initial = ylable
                                    # 插入6u补空板
                                    ylable = ylable_initial - 4 * one_u_high
                                    insertBlockFun(xlable_initial, ylable, dwg_address, '4U补空板')
                                    # 更新y轴坐标
                                    ylable_initial = ylable
                                else:
                                    for k in range(high_value):
                                        ylable = ylable_initial - one_u_high
                                        # 插入1u补空板
                                        insertBlockFun(xlable_initial, ylable, dwg_address, '1U补空板')
                                        # 更新y轴坐标
                                        ylable_initial = ylable
                            else:
                                print('"尺寸"列为空，跳过补空板生成！')
                    except:
                        continue
                print('{0} 生成完毕'.format(execute_sheet_name))
        else:
            print('执行机柜配置sheet不存在!')
        # 保存文件
        print('DWG自动生成完毕，另存为：Model_SaveAs.dwg，在当前文件下！')

        add_hour = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()).split()[1].replace(':', '_')
        acad.ActiveDocument.SaveAs(current_path+'\\DWG_SaveAs_{0}_{1}.dwg'.format(station_name, add_hour))
        # 生成设备清单
        # 先将ips_jigui_dic,execute_jigui_dic 两个list中的内容进行组合，然后传递到设备清单函数中
        print(ips_jigui_dic, execute_jigui_dic)
        equipmentsListExecl(ips_jigui_dic, execute_jigui_dic, station_name)

    else:
        print('联锁机机柜配置sheet和执行机柜配置sheet都不存在！！')
    # 是否有‘联锁机机柜配置’sheet，如果有就画这个图，没有就不画

else:
    print('DWG文件夹或Config.xlsx文件不存在，程序自动退出')
    print('==============================================================')
